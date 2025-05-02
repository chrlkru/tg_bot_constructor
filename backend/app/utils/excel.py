# app/utils/excel.py

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def append_rows(
    file_path: Path | str,
    headers: list[str],
    rows: list[list],
) -> Path:
    """
    Append rows of data to an Excel file.
    - If the file does not exist, creates it and writes the headers first.
    - Adjusts column widths to fit content.
    Returns the Path to the Excel file.
    """
    file_path = Path(file_path)
    if file_path.exists():
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    for row in rows:
        ws.append(row)

    # Авто-ширина колонок
    for idx, col_cells in enumerate(ws.columns, start=1):
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

    wb.save(file_path)
    return file_path

def append_survey_result(
    file_path: Path | str,
    survey_id: int,
    user_id: int,
    answers: dict[int, str],
    question_order: list[int],
) -> Path:
    """
    Specialization for quiz/ survey results:
    - Headers: ['timestamp','survey_id','user_id'] + ['q_<id>'...]
    - One row per submission with timestamp, survey_id, user_id, then answers in question_order.
    """
    timestamp = datetime.now().isoformat()
    headers = ['timestamp', 'survey_id', 'user_id'] + [f"q_{qid}" for qid in question_order]
    row = [timestamp, survey_id, user_id] + [answers.get(qid, "") for qid in question_order]
    return append_rows(file_path, headers, [row])

async def send_excel_report(
    bot,
    chat_id: int,
    file_path: Path | str,
    caption: str = "",
):
    """
    Sends the given Excel file to a Telegram chat.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    with path.open("rb") as f:
        await bot.send_document(chat_id, f, caption=caption)
